"""Import validated sales documents exported from the external ERP system."""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any

import frappe
from frappe import _
from frappe.utils import flt, getdate, nowdate

from fmcg_wms.services.sales_order import IMMEDIATE_DELIVERY_MODE

SALES_ORDER_IMPORT = "\u9500\u552e\u8ba2\u5355"
DELIVERY_NOTE_IMPORT = "\u9500\u8d27\u5355"
CUSTOMER_IMPORT = "\u5ba2\u6237\u57fa\u7840\u8d44\u6599"
ITEM_IMPORT = "\u7269\u6599\u57fa\u7840\u8d44\u6599"

ORDER_REQUIRED_COLUMNS = {"\u5355\u636e\u7f16\u53f7", "\u5ba2\u6237\u7f16\u7801", "\u5b58\u8d27\u7f16\u7801", "\u6570\u91cf", "\u9500\u552e\u5355\u4f4d", "\u4ed3\u5e93\u7f16\u7801"}
DELIVERY_REQUIRED_COLUMNS = ORDER_REQUIRED_COLUMNS | {"\u9500\u552e\u8ba2\u5355\u53f7"}
CUSTOMER_REQUIRED_COLUMNS = {"\u5f80\u6765\u5355\u4f4d\u7f16\u7801", "\u5f80\u6765\u5355\u4f4d\u540d\u79f0", "\u6027\u8d28", "\u505c\u7528"}
ITEM_REQUIRED_COLUMNS = {"\u5b58\u8d27\u7f16\u7801", "\u5b58\u8d27\u540d\u79f0", "\u6240\u5c5e\u7c7b\u522b", "\u54c1\u724c", "\u8ba1\u91cf\u5355\u4f4d"}


def preview_import(
    file_url: str,
    import_type: str,
    company: str,
    customer_group: str | None = None,
    territory: str | None = None,
    item_group: str | None = None,
    sales_uom: str | None = None,
) -> dict:
    """Parse and validate an uploaded file without creating any document."""
    frappe.only_for("System Manager")
    rows = _read_rows(file_url, import_type)
    plans, errors = _build_import_plan(rows, import_type, company, customer_group, territory, item_group, sales_uom)
    return _result(rows, plans, errors, preview=True)


def run_import(
    file_url: str,
    import_type: str,
    company: str,
    submit_documents: bool = False,
    customer_group: str | None = None,
    territory: str | None = None,
    item_group: str | None = None,
    sales_uom: str | None = None,
) -> dict:
    """Create external ERP documents only after the complete file passes validation."""
    frappe.only_for("System Manager")
    rows = _read_rows(file_url, import_type)
    plans, errors = _build_import_plan(rows, import_type, company, customer_group, territory, item_group, sales_uom)
    if errors:
        return _result(rows, plans, errors, preview=False)

    created = []
    for plan in plans:
        if plan["skip"]:
            continue
        if import_type == SALES_ORDER_IMPORT:
            doc = _create_sales_order(plan, submit_documents)
        elif import_type == DELIVERY_NOTE_IMPORT:
            doc = _create_delivery_note(plan, submit_documents)
        elif import_type == CUSTOMER_IMPORT:
            doc = _create_customer(plan)
        else:
            doc = _create_item(plan)
        created.append(doc.name)

    result = _result(rows, plans, [], preview=False)
    result["created_documents"] = created
    return result


def _read_rows(file_url: str, import_type: str) -> list[dict[str, Any]]:
    required_columns_by_type = {
        SALES_ORDER_IMPORT: ORDER_REQUIRED_COLUMNS,
        DELIVERY_NOTE_IMPORT: DELIVERY_REQUIRED_COLUMNS,
        CUSTOMER_IMPORT: CUSTOMER_REQUIRED_COLUMNS,
        ITEM_IMPORT: ITEM_REQUIRED_COLUMNS,
    }
    if import_type not in required_columns_by_type:
        frappe.throw(_("Unsupported import type."))
    required_columns = required_columns_by_type[import_type]

    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        frappe.throw(_("Upload the Excel file to the import document before previewing it."))
    file_doc = frappe.get_doc("File", file_name)
    file_doc.check_permission("read")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        frappe.throw(_("The server is missing the Excel import dependency: {0}").format(exc))

    workbook = load_workbook(filename=BytesIO(file_doc.get_content()), read_only=True, data_only=True)
    worksheet = workbook.active
    headers, header_row = _find_headers(worksheet, required_columns)
    rows = []
    for excel_row, values in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        row = {header: _cell_value(values[index]) for header, index in headers.items()}
        if _row_has_import_key(row, import_type):
            row["_row_number"] = excel_row
            rows.append(row)
    if not rows:
        frappe.throw(_("No document detail rows were found in the Excel file."))
    return rows


def _find_headers(worksheet, required_columns: set[str]) -> tuple[dict[str, int], int]:
    for row_number, cells in enumerate(worksheet.iter_rows(), start=1):
        headers = {_text(cell.value): index for index, cell in enumerate(cells) if _text(cell.value)}
        if required_columns.issubset(headers):
            return headers, row_number
    frappe.throw(
        _("Excel is missing required columns: {0}").format(", ".join(sorted(required_columns)))
    )


def _cell_value(cell):
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, (int, float)) and re.fullmatch(r"0+", cell.number_format or ""):
        return str(int(value)).zfill(len(cell.number_format))
    return value


def _build_import_plan(
    rows: list[dict],
    import_type: str,
    company: str,
    customer_group: str | None,
    territory: str | None,
    item_group: str | None,
    sales_uom: str | None,
) -> tuple[list[dict], list[str]]:
    if not frappe.db.exists("Company", company):
        return [], [_('Company "{0}" does not exist.').format(company)]
    _require_external_fields()
    if import_type == CUSTOMER_IMPORT:
        customer_group = _default_customer_group(customer_group)
        territory = _default_territory(territory)
    key_field = {
        SALES_ORDER_IMPORT: "\u5355\u636e\u7f16\u53f7",
        DELIVERY_NOTE_IMPORT: "\u5355\u636e\u7f16\u53f7",
        CUSTOMER_IMPORT: "\u5f80\u6765\u5355\u4f4d\u7f16\u7801",
        ITEM_IMPORT: "\u5b58\u8d27\u7f16\u7801",
    }[import_type]
    groups = _group_rows(rows, key_field)
    plans, errors = [], []
    for external_document_no, document_rows in groups.items():
        try:
            if import_type == SALES_ORDER_IMPORT:
                plan = _plan_sales_order(external_document_no, document_rows, company)
            elif import_type == DELIVERY_NOTE_IMPORT:
                plan = _plan_delivery_note(external_document_no, document_rows, company)
            elif import_type == CUSTOMER_IMPORT:
                plan = _plan_customer(external_document_no, document_rows, customer_group, territory)
            else:
                plan = _plan_item(external_document_no, document_rows, item_group, sales_uom)
            plans.append(plan)
        except frappe.ValidationError as exc:
            errors.append(_document_error(external_document_no, str(exc)))
    return plans, errors


def _plan_customer(
    external_code: str, rows: list[dict], customer_group: str | None, territory: str | None
) -> dict:
    if len(rows) != 1:
        frappe.throw(_("Customer code {0} appears more than once in the Excel file.").format(external_code))
    row = rows[0]
    relationship_type = _text(row.get("\u6027\u8d28"))
    if "\u5ba2\u6237" not in relationship_type or _text(row.get("\u505c\u7528")) in {"\u662f", "yes", "y", "true", "1"}:
        return {
            "external_no": external_code,
            "rows": rows,
            "skip": True,
            "existing": _("Filtered: not an active customer relationship"),
        }
    if frappe.db.exists("Customer", {"fmcg_external_customer_code": external_code}):
        existing = frappe.db.get_value("Customer", {"fmcg_external_customer_code": external_code}, "name")
        return {"external_no": external_code, "rows": rows, "skip": True, "existing": existing}
    _require_link("Customer Group", customer_group, _("Customer Group"))
    _require_link("Territory", territory, _("Territory"))
    return {
        "external_no": external_code,
        "rows": rows,
        "skip": False,
        "customer_name": _required_value(row, "\u5f80\u6765\u5355\u4f4d\u540d\u79f0"),
        "customer_group": customer_group,
        "territory": territory,
    }


def _plan_item(external_code: str, rows: list[dict], item_group: str | None, sales_uom: str | None) -> dict:
    if len(rows) != 1:
        frappe.throw(_("Item code {0} appears more than once in the Excel file.").format(external_code))
    row = rows[0]
    existing = frappe.db.get_value("Item", {"fmcg_external_item_code": external_code}, "name")
    if existing:
        return {"external_no": external_code, "rows": rows, "skip": True, "existing": existing}
    _require_link("Item Group", item_group, _("Item Group"))
    _require_link("UOM", sales_uom, _("Sales UOM"))
    is_freight = external_code.upper() == "FREIGHT"
    packaging = _parse_packaging(row) if not is_freight else {"stock_uom": sales_uom, "conversion_factor": 1}
    _require_link("UOM", packaging["stock_uom"], _("Stock UOM"))
    return {
        "external_no": external_code,
        "rows": rows,
        "skip": False,
        "item_name": _required_value(row, "\u5b58\u8d27\u540d\u79f0"),
        "item_group": item_group,
        "stock_uom": packaging["stock_uom"],
        "sales_uom": sales_uom,
        "conversion_factor": packaging["conversion_factor"],
        "description": _item_description(row),
        "is_stock_item": 0 if is_freight else 1,
    }


def _plan_sales_order(external_order_no: str, rows: list[dict], company: str) -> dict:
    existing_name = frappe.db.get_value(
        "Sales Order", {"fmcg_external_order_no": external_order_no, "docstatus": ["!=", 2]}, "name"
    )
    if existing_name:
        return {"external_no": external_order_no, "rows": rows, "skip": True, "existing": existing_name}

    first = rows[0]
    customer = _resolve_customer(first)
    warehouse = _resolve_warehouse(first, company)
    mode = _customer_delivery_mode(customer)
    item_lines, line_keys = [], set()
    for row in rows:
        if _resolve_customer(row) != customer:
            frappe.throw(_("Rows for one order must use the same customer."))
        row_warehouse = _resolve_warehouse(row, company)
        item_code = _resolve_item(row)
        line_key = _line_key(external_order_no, row)
        if line_key in line_keys:
            frappe.throw(
                _("Duplicate item detail key at Excel row {0}. Add an external line number to distinguish it.").format(
                    row["_row_number"]
                )
            )
        line_keys.add(line_key)
        item_lines.append(
            {
                "item_code": item_code,
                "qty": _positive_qty(row),
                "uom": _required_value(row, "\u9500\u552e\u5355\u4f4d"),
                "warehouse": row_warehouse,
                "rate": _rate(row),
                "is_free_item": _is_gift(row),
                "external_line_key": line_key,
            }
        )
    return {
        "external_no": external_order_no,
        "rows": rows,
        "skip": False,
        "company": company,
        "customer": customer,
        "warehouse": warehouse,
        "delivery_mode": mode,
        "transaction_date": _document_date(first, "\u5355\u636e\u65e5\u671f"),
        "delivery_date": _document_date(first, "\u4ea4\u8d27\u65e5\u671f"),
        "items": item_lines,
    }


def _plan_delivery_note(external_delivery_no: str, rows: list[dict], company: str) -> dict:
    existing_name = frappe.db.get_value(
        "Delivery Note", {"fmcg_external_delivery_no": external_delivery_no, "docstatus": ["!=", 2]}, "name"
    )
    if existing_name:
        return {"external_no": external_delivery_no, "rows": rows, "skip": True, "existing": existing_name}

    planned_lines, line_keys, customer = [], set(), None
    for row in rows:
        external_order_no = _required_value(row, "\u9500\u552e\u8ba2\u5355\u53f7")
        sales_order_name = frappe.db.get_value(
            "Sales Order",
            {"fmcg_external_order_no": external_order_no, "docstatus": 1},
            "name",
        )
        if not sales_order_name:
            frappe.throw(
                _("External Sales Order {0} is not imported and submitted.").format(external_order_no)
            )
        sales_order = frappe.get_doc("Sales Order", sales_order_name)
        if sales_order.company != company:
            frappe.throw(_("Sales Order {0} belongs to another company.").format(sales_order.name))
        row_customer = _resolve_customer(row)
        if row_customer != sales_order.customer:
            frappe.throw(_("Customer does not match Sales Order {0}.").format(sales_order.name))
        if customer and customer != row_customer:
            frappe.throw(_("Rows for one delivery note must use the same customer."))
        customer = row_customer
        line_key = _line_key(external_order_no, row)
        if line_key in line_keys:
            frappe.throw(
                _("Duplicate delivery item detail key at Excel row {0}.").format(row["_row_number"])
            )
        line_keys.add(line_key)
        matching_rows = [item for item in sales_order.items if item.fmcg_external_line_key == line_key]
        if len(matching_rows) != 1:
            frappe.throw(
                _("Could not find one Sales Order Item for external key {0}.").format(line_key)
            )
        planned_lines.append(
            {
                "sales_order": sales_order,
                "sales_order_item": matching_rows[0],
                "qty": _positive_qty(row),
                "warehouse": _resolve_warehouse(row, company),
                "external_line_key": line_key,
            }
        )
    return {
        "external_no": external_delivery_no,
        "rows": rows,
        "skip": False,
        "company": company,
        "customer": customer,
        "posting_date": _document_date(rows[0], "\u9001\u8d27\u65e5\u671f", fallback="\u5355\u636e\u65e5\u671f"),
        "items": planned_lines,
    }


def _create_sales_order(plan: dict, submit_documents: bool):
    sales_order = frappe.new_doc("Sales Order")
    sales_order.company = plan["company"]
    sales_order.customer = plan["customer"]
    sales_order.transaction_date = plan["transaction_date"]
    sales_order.delivery_date = plan["delivery_date"]
    sales_order.set_warehouse = plan["warehouse"]
    sales_order.fmcg_delivery_mode = plan["delivery_mode"]
    sales_order.fmcg_external_order_no = plan["external_no"]
    sales_order.run_method("set_missing_values")

    item_meta = frappe.get_meta("Sales Order Item")
    for line in plan["items"]:
        item = {
            "item_code": line["item_code"],
            "qty": line["qty"],
            "uom": line["uom"],
            "warehouse": line["warehouse"],
            "rate": 0 if line["is_free_item"] else line["rate"],
            "fmcg_external_line_key": line["external_line_key"],
        }
        if item_meta.has_field("is_free_item"):
            item["is_free_item"] = 1 if line["is_free_item"] else 0
        sales_order.append("items", item)
    sales_order.run_method("calculate_taxes_and_totals")
    sales_order.insert()
    if submit_documents:
        sales_order.submit()
    return sales_order


def _create_delivery_note(plan: dict, submit_documents: bool):
    from erpnext.selling.doctype.sales_order.sales_order import make_delivery_note

    delivery_note = frappe.new_doc("Delivery Note")
    delivery_note.company = plan["company"]
    delivery_note.customer = plan["customer"]
    delivery_note.posting_date = plan["posting_date"]
    delivery_note.fmcg_external_delivery_no = plan["external_no"]
    mapped_notes = {}
    for line in plan["items"]:
        sales_order = line["sales_order"]
        if sales_order.name not in mapped_notes:
            mapped_notes[sales_order.name] = make_delivery_note(sales_order.name)
        mapped_item = next(
            (item for item in mapped_notes[sales_order.name].items if item.so_detail == line["sales_order_item"].name),
            None,
        )
        if not mapped_item:
            frappe.throw(_("Could not map Sales Order Item {0} to a Delivery Note.").format(line["sales_order_item"].name))
        delivery_note.append(
            "items",
            {
                "item_code": mapped_item.item_code,
                "item_name": mapped_item.item_name,
                "description": mapped_item.description,
                "qty": line["qty"],
                "stock_qty": line["qty"] * flt(mapped_item.conversion_factor or 1),
                "uom": mapped_item.uom,
                "stock_uom": mapped_item.stock_uom,
                "conversion_factor": mapped_item.conversion_factor or 1,
                "warehouse": line["warehouse"],
                "rate": mapped_item.rate,
                "price_list_rate": mapped_item.price_list_rate,
                "against_sales_order": sales_order.name,
                "so_detail": line["sales_order_item"].name,
                "fmcg_external_line_key": line["external_line_key"],
            },
        )
    delivery_note.set_warehouse = delivery_note.items[0].warehouse
    delivery_note.run_method("set_missing_values")
    delivery_note.run_method("set_po_nos")
    delivery_note.run_method("calculate_taxes_and_totals")
    delivery_note.insert()
    if submit_documents:
        delivery_note.submit()
    return delivery_note


def _create_customer(plan: dict):
    customer = frappe.new_doc("Customer")
    customer.customer_name = plan["customer_name"]
    customer.customer_type = "Company"
    customer.customer_group = plan["customer_group"]
    customer.territory = plan["territory"]
    customer.fmcg_external_customer_code = plan["external_no"]
    customer.insert()
    return customer


def _create_item(plan: dict):
    item = frappe.new_doc("Item")
    item.item_name = plan["item_name"]
    item.item_group = plan["item_group"]
    item.stock_uom = plan["stock_uom"]
    item.sales_uom = plan["sales_uom"]
    item.is_stock_item = plan["is_stock_item"]
    item.description = plan["description"]
    item.fmcg_external_item_code = plan["external_no"]
    if plan["sales_uom"] != plan["stock_uom"]:
        item.append(
            "uoms",
            {"uom": plan["sales_uom"], "conversion_factor": plan["conversion_factor"]},
        )
    # Let ERPNext apply its configured Item naming rule instead of using the external code.
    item.run_method("set_missing_values")
    item.insert()
    return item


def _resolve_customer(row: dict) -> str:
    external_code = _required_value(row, "\u5ba2\u6237\u7f16\u7801")
    customer_name = _text(row.get("\u5ba2\u6237"))
    matches = frappe.get_all(
        "Customer", filters={"fmcg_external_customer_code": external_code, "disabled": 0}, pluck="name"
    )
    if not matches and frappe.db.exists("Customer", external_code):
        matches = [external_code]
    if not matches and customer_name:
        matches = frappe.get_all("Customer", filters={"customer_name": customer_name, "disabled": 0}, pluck="name")
    if len(matches) != 1:
        frappe.throw(_("Customer code {0} needs one ERPNext customer mapping.").format(external_code))
    return matches[0]


def _resolve_item(row: dict) -> str:
    external_code = _required_value(row, "\u5b58\u8d27\u7f16\u7801")
    matches = frappe.get_all(
        "Item", filters={"fmcg_external_item_code": external_code, "disabled": 0}, pluck="name"
    )
    if not matches and frappe.db.exists("Item", external_code):
        matches = [external_code]
    if len(matches) != 1:
        frappe.throw(_("Item code {0} needs one ERPNext item mapping.").format(external_code))
    return matches[0]


def _resolve_warehouse(row: dict, company: str) -> str:
    external_code = _required_value(row, "\u4ed3\u5e93\u7f16\u7801")
    matches = frappe.get_all(
        "Warehouse",
        filters={"fmcg_external_warehouse_code": external_code, "company": company, "is_group": 0},
        pluck="name",
    )
    if not matches:
        matches = frappe.get_all(
            "Warehouse", filters={"name": external_code, "company": company, "is_group": 0}, pluck="name"
        )
    if len(matches) != 1:
        frappe.throw(_("Warehouse code {0} needs one ERPNext warehouse mapping.").format(external_code))
    return matches[0]


def _customer_delivery_mode(customer: str) -> str:
    return frappe.db.get_value("Customer", customer, "fmcg_default_delivery_mode") or IMMEDIATE_DELIVERY_MODE


def _line_key(external_order_no: str, row: dict) -> str:
    return "|".join(
        [external_order_no, _required_value(row, "\u5b58\u8d27\u7f16\u7801"), "gift" if _is_gift(row) else "sale"]
    )


def _is_gift(row: dict) -> bool:
    return _text(row.get("\u8d60\u54c1")).lower() in {"\u662f", "yes", "y", "true", "1"}


def _positive_qty(row: dict) -> float:
    qty = flt(row.get("\u6570\u91cf"))
    if qty <= 0:
        frappe.throw(
            _("Excel row {0} has a non-positive quantity. Import returns separately, not as a sales delivery.").format(
                row["_row_number"]
            )
        )
    return qty


def _rate(row: dict) -> float:
    inclusive_rate = flt(row.get("\u542b\u7a0e\u5355\u4ef7"))
    return inclusive_rate or flt(row.get("\u5355\u4ef7"))


def _document_date(row: dict, fieldname: str, fallback: str | None = None):
    value = row.get(fieldname) or (row.get(fallback) if fallback else None)
    return getdate(value) if value else getdate(nowdate())


def _required_value(row: dict, fieldname: str) -> str:
    value = _text(row.get(fieldname))
    if not value:
        frappe.throw(_("Excel row {0} is missing {1}.").format(row.get("_row_number"), fieldname))
    return value


def _group_rows(rows: list[dict], fieldname: str) -> dict[str, list[dict]]:
    groups = defaultdict(list)
    for row in rows:
        groups[_required_value(row, fieldname)].append(row)
    return groups


def _require_external_fields() -> None:
    required = {
        "Customer": ["fmcg_default_delivery_mode", "fmcg_external_customer_code"],
        "Item": ["fmcg_external_item_code"],
        "Warehouse": ["fmcg_external_warehouse_code"],
        "Sales Order": ["fmcg_external_order_no"],
        "Sales Order Item": ["fmcg_external_line_key"],
        "Delivery Note": ["fmcg_external_delivery_no"],
        "Delivery Note Item": ["fmcg_external_line_key"],
    }
    missing = [
        f"{doctype}.{fieldname}"
        for doctype, fields in required.items()
        for fieldname in fields
        if not frappe.get_meta(doctype).has_field(fieldname)
    ]
    if missing:
        frappe.throw(_("Run bench migrate before using external ERP import: {0}").format(", ".join(missing)))


def _result(rows: list[dict], plans: list[dict], errors: list[str], preview: bool) -> dict:
    skipped = [plan for plan in plans if plan["skip"]]
    ready = [plan for plan in plans if not plan["skip"]]
    return {
        "preview": preview,
        "source_rows": len(rows),
        "ready_documents": len(ready),
        "skipped_documents": len(skipped),
        "skipped": [{"external_no": plan["external_no"], "existing": plan["existing"]} for plan in skipped],
        "errors": errors,
    }


def _document_error(external_document_no: str, message: str) -> str:
    return _("External document {0}: {1}").format(external_document_no, message)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_has_import_key(row: dict, import_type: str) -> bool:
    fieldname = {
        SALES_ORDER_IMPORT: "\u5355\u636e\u7f16\u53f7",
        DELIVERY_NOTE_IMPORT: "\u5355\u636e\u7f16\u53f7",
        CUSTOMER_IMPORT: "\u5f80\u6765\u5355\u4f4d\u7f16\u7801",
        ITEM_IMPORT: "\u5b58\u8d27\u7f16\u7801",
    }[import_type]
    return bool(_text(row.get(fieldname)))


def _require_link(doctype: str, value: str | None, label: str) -> None:
    if not value or not frappe.db.exists(doctype, value):
        frappe.throw(_("Select an existing {0} before importing.").format(label))


def _default_customer_group(value: str | None) -> str:
    return _default_master_value("Customer Group", value, "\u4e2a\u4eba", _("Customer Group"))


def _default_territory(value: str | None) -> str:
    return _default_master_value("Territory", value, "\u4e2d\u56fd", _("Territory"))


def _default_master_value(doctype: str, value: str | None, preferred: str, label: str) -> str:
    if value and frappe.db.exists(doctype, value):
        return value
    if frappe.db.exists(doctype, preferred):
        return preferred
    candidate = frappe.db.get_value(doctype, {"is_group": 0}, "name", order_by="lft")
    if not candidate:
        candidate = frappe.db.get_value(doctype, {}, "name", order_by="name")
    if not candidate:
        frappe.throw(_("Create one {0} before importing customers.").format(label))
    return candidate


def _item_description(row: dict) -> str:
    details = [
        (_("Specification"), _text(row.get("\u89c4\u683c\u578b\u53f7"))),
        (_("Source Category"), _text(row.get("\u6240\u5c5e\u7c7b\u522b"))),
        (_("Brand"), _text(row.get("\u54c1\u724c"))),
        (_("Source UOM"), _text(row.get("\u8ba1\u91cf\u5355\u4f4d"))),
    ]
    return "\n".join(f"{label}: {value}" for label, value in details if value)


def _parse_packaging(row: dict) -> dict:
    source_uom = _required_value(row, "\u8ba1\u91cf\u5355\u4f4d")
    match = re.fullmatch(r"1\s*(?:\u7bb1|\u4ef6)\s*=\s*(\d+(?:\.\d+)?)\s*([\u4e2a\u74f6\u7f50\u542c\u7ec4])", source_uom)
    if not match:
        frappe.throw(
            _("Excel row {0} has an unsupported packaging format {1}. Use a value such as 1\u7bb1=6\u74f6, 1\u7bb1=24\u7f50, 1\u7bb1=12\u4e2a, or 1\u7bb1=2\u7ec4.").format(
                row["_row_number"], source_uom
            )
        )
    source_base_uom = match.group(2)
    stock_uom = "\u7f50" if source_base_uom == "\u542c" else source_base_uom
    return {"stock_uom": stock_uom, "conversion_factor": flt(match.group(1))}
